from django.shortcuts import render, redirect, get_object_or_404
from .models import Servidores
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
from django.utils import timezone
from django.contrib import messages
from .models import Nodos,ApisYsubdominios
from .forms import NodosForm
from .forms import FormularioForm, Formulario, SubdominioForm, Subdominios
from .forms import MyModelForm, CustomUserCreationForm,ApisYsubdominiosForm
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import openpyxl
from django.http import HttpResponse
from .models import ApisYsubdominios

def home(request):
    return render(request, 'registration/login.html')

def register(request):
    data = {'form': CustomUserCreationForm()}
    if request.method == 'POST':
        user_creation_form = CustomUserCreationForm(data=request.POST)
        if user_creation_form.is_valid():
            user = user_creation_form.save()
            user = authenticate(username=user_creation_form.cleaned_data['username'], password=user_creation_form.cleaned_data['password1'])
            if user is not None:
                login(request, user)
                messages.success(request, '¡Usuario registrado correctamente!')
                return redirect('register')
    return render(request, 'registration/register.html', data)

@login_required
@permission_required('myapp.view_mymodel', raise_exception=True)
def model_detail(request, pk):
    model = get_object_or_404(Servidores, pk=pk)
    context = {
        'user': request.user.get_full_name(),
        'model': model
    }
    return render(request, 'model_detail.html', {'model': model})

@login_required
@permission_required('myapp.view_mymodel', raise_exception=True)
def model_list(request):
    models = Servidores.objects.all()
    nodos = Nodos.objects.all()
    apis = ApisYsubdominios.objects.all()
    formulario = Formulario.objects.all()
    subdominios = Subdominios.objects.all()

    return render(request, 'model_list.html', {
        'models': models,
        'nodos': nodos,
        'apis': apis,
        'formulario': formulario,
        'subdominios': subdominios


    })

@login_required
@permission_required('myapp.view_mymodel', raise_exception=True)
def model_create(request):
    if request.method == 'POST':
        form = MyModelForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            for field_name, field_value in instance.__dict__.items():
                if not field_value and field_name != 'id':
                    setattr(instance, field_name, 'VACÍO')
            instance.save()
            return redirect('model_list')
    else:
        form = MyModelForm()
    return render(request, 'model_form.html', {'form': form})


@login_required
def model_update(request, pk):
    model = get_object_or_404(Servidores, pk=pk)
    if request.method == 'POST':
        form = MyModelForm(request.POST, request.FILES, instance=model)
        if form.is_valid():
            model = form.save(commit=False)
            if 'archivo' in request.FILES:
                model.archivo = request.FILES['archivo']
            model.save()
            return redirect('model_list')  # Redirigir a la lista de modelos después de la actualización
    else:
        form = MyModelForm(instance=model)

    return render(request, 'actualizar.html', {'form': form, 'model': model})


@login_required
@permission_required('myapp.view_mymodel', raise_exception=True)
def model_confirm_actualizar(request, pk):
    model = get_object_or_404(Servidores, pk=pk)
    if request.method == 'POST':
        form = MyModelForm(request.POST, instance=model)
        if form.is_valid():
            form.save()
            return redirect('model_list')
    else:
        form = MyModelForm(instance=model)
    return render(request, 'model_confirm_actualizar.html', {'form': form, 'model': model})

@login_required
@permission_required('myapp.can_view_mymodel', raise_exception=True)
def model_delete(request, pk):
    model = get_object_or_404(Servidores, pk=pk)
    if request.method == 'POST':
        model.estado_registro = False
        model.delete()
        return redirect('model_list')
    return redirect('model_confirm_delete', pk=pk)

@login_required
@permission_required('myapp.view_mymodel', raise_exception=True)
def model_confirm_delete(request, pk):
    model = get_object_or_404(Servidores, pk=pk)
    return render(request, 'model_confirm_delete.html', {'model': model})

@login_required
def salir(request):
    logout(request)
    return redirect('/')

# Exportación a Excel
def export_to_excel_servidores(request):
    queryset = Servidores.objects.all()

    # Crear un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    title = 'Reporte de Servidores'
    ws.append([title])

    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=25, bold=True)  # Negrita y tamaño de fuente
    ws.merge_cells('A1:I1')  # Combinar celdas para el título
    title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto horizontal y verticalmente
    total_columns = 9 
    additional_info = [
        'INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI',
        'BIENES POR CUSTODIO',
        'REGISTRO DE CONSTATACIÓN FÍSICA DE BIENES 2024'
    ]
    start_row_for_info = 2
    for i, info in enumerate(additional_info, start=start_row_for_info):
        # Añadir una fila en blanco para que la información adicional se pueda colocar en varias celdas
        ws.append([''] * total_columns)
        info_cell = ws.cell(row=i, column=1)
        info_cell.value = info
        info_cell.font = Font(size=14, bold=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)
    ws.append([''] * total_columns)

    # Escribir encabezados de columna
    column_names = ['ID','DireccionIp', 'Usuario', 'Contrasena', 'Servicio', 'Puerto', 'RutaImportante', 
                    'UbicacionFisica', 'NumeroSerie']
    ws.append(column_names)

    for cell in ws[ws.max_row]:  # Los encabezados están en la fila 3
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    # Escribir datos de los objetos en el libro
    for obj in queryset:
        ws.append([obj.id, obj.DireccionIp, obj.Usuario, obj.Contrasena, obj.Servicio, obj.Puerto, obj.RutaImportante, obj.UbicacionFisica, 
                   obj.NumeroSerie])
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=ws.max_row):  # Empezando desde la fila 4
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto
    
    column_widths = [10, 30, 20, 20, 30, 20, 20, 20, 20]  # Ajustar estos valores según tus necesidades
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  # A, B, C, D, E corresponden a 1, 2, 3, 4, 5

    
    # Crear una respuesta de HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Servidores_Report.xlsx'
    wb.save(response)
    return response
#Nodo ------------------------------------------------------------------------
@login_required
@permission_required('myapp.view_nodos', raise_exception=True)
def nodos_list(request):
    nodos = Nodos.objects.all()
    return render(request, 'model_list.html', {'nodos': nodos})

@login_required
@permission_required('myapp.add_nodos', raise_exception=True)
def nodos_create(request):
    if request.method == 'POST':
        form = NodosForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Nodo creado correctamente!')
            return redirect('model_list')
    else:
        form = NodosForm()
    return render(request, 'nodos/nodos_form.html', {'form': form})

@login_required
@permission_required('myapp.change_nodos', raise_exception=True)
def nodos_update(request, pk):
    nodo = get_object_or_404(Nodos, pk=pk)
    if request.method == 'POST':
        form = NodosForm(request.POST, instance=nodo)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Nodo actualizado correctamente!')
            return redirect('model_list')
    else:
        form = NodosForm(instance=nodo)
    return render(request, 'nodos/nodos_form.html', {'form': form, 'nodo': nodo})

@login_required
@permission_required('myapp.delete_nodos', raise_exception=True)
def nodos_delete(request, pk):
    nodo = get_object_or_404(Nodos, pk=pk)
    if request.method == 'POST':
        nodo.delete()
        messages.success(request, '¡Nodo eliminado correctamente!')
        return redirect('model_list')
    return render(request, 'nodos/nodos_confirm_delete.html', {'nodo': nodo})

@login_required
@permission_required('myapp.view_nodos', raise_exception=True)
def nodos_detail(request, pk):
    nodo = get_object_or_404(Nodos, pk=pk)
    return render(request, 'nodos/nodos_detail.html', {'nodo': nodo})


# Exportación a Excel NODOS
def export_nodos_to_excel(request):
    queryset = Nodos.objects.all()

    # Crear un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    
    title = 'Reporte de Nodos'
    ws.append([title])

    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=25, bold=True)  # Negrita y tamaño de fuente
    ws.merge_cells('A1:H1')  # Combinar celdas para el título
    title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto horizontal y verticalmente
    total_columns = 8 

    additional_info = [
        'INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI',
        'BIENES POR CUSTODIO',
        'REGISTRO DE CONSTATACIÓN FÍSICA DE BIENES 2024'
    ]
    start_row_for_info = 2
    for i, info in enumerate(additional_info, start=start_row_for_info):
        # Añadir una fila en blanco para que la información adicional se pueda colocar en varias celdas
        ws.append([''] * total_columns)
        info_cell = ws.cell(row=i, column=1)
        info_cell.value = info
        info_cell.font = Font(size=14, bold=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)
    ws.append([''] * total_columns)
    # Escribir encabezados de columna
    column_names = ['ID','Host', 'Usuario', 'Ram', 'Disco', 'SistemaOperativo', 'Descripcion', 'Contrasena']
    ws.append(column_names)

     # Aplicar formato a los encabezados de columna
    for cell in ws[ws.max_row]:  # Los encabezados están en la fila 3
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

   
    # Escribir datos de los objetos en el libro
    for obj in queryset:
        ws.append([obj.id, obj.Host, obj.Usuario, obj.Ram, obj.Disco, obj.SistemaOperativo, obj.Descripcion, obj.Contrasena])

    for row in ws.iter_rows(min_row=4, max_col=5, max_row=ws.max_row):  # Empezando desde la fila 4
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    column_widths = [10, 30, 10, 10, 10, 30, 40, 30]  # Ajustar estos valores según tus necesidades
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  # A, B, C, D, E corresponden a 1, 2, 3, 4, 5

    # Crear una respuesta de HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Nodos_Report.xlsx'
    wb.save(response)
    return response

# APIS -----------------------------------------------------------------------


@login_required
@permission_required('myapp.view_apisysubdominios', raise_exception=True)
def apis_view(request):
    apis = ApisYsubdominios.objects.all()
    return render(request, 'model_list.html', {'apis': apis})

@login_required
@permission_required('myapp.add_apisysubdominios', raise_exception=True)
def apis_create(request):
    if request.method == 'POST':
        form = ApisYsubdominiosForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡API creada correctamente!')
            return redirect('model_list')
    else:
        form = ApisYsubdominiosForm()
    return render(request, 'crear_apis.html', {'form': form})

@login_required
@permission_required('myapp.change_apisysubdominios', raise_exception=True)
def apis_update(request, pk):
    api = get_object_or_404(ApisYsubdominios, pk=pk)
    if request.method == 'POST':
        form = ApisYsubdominiosForm(request.POST, instance=api)
        if form.is_valid():
            form.save()
            messages.success(request, '¡API actualizada correctamente!')
            return redirect('model_list')
    else:
        form = ApisYsubdominiosForm(instance=api)
    return render(request, 'apis_update.html', {'form': form, 'api': api})

@login_required
@permission_required('myapp.delete_apisysubdominios', raise_exception=True)
def apis_delete(request, pk):
    api = get_object_or_404(ApisYsubdominios, pk=pk)
    if request.method == 'POST':
        api.delete()
        messages.success(request, '¡API eliminada correctamente!')
        return redirect('model_list')
    return render(request, 'apis_confirm_delete.html', {'api': api})

@login_required
@permission_required('myapp.view_apisysubdominios', raise_exception=True)
def apis_detail(request, pk):
    api = get_object_or_404(ApisYsubdominios, pk=pk)
    return render(request, 'apis_detail.html', {'api': api})


def export_Api_to_excel(request):
    queryset = ApisYsubdominios.objects.all()

    wb = Workbook()
    ws = wb.active

    title = 'Reporte de Apis'
    ws.append([title])

    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=25, bold=True)
    ws.merge_cells('A1:E1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_columns = 5

    additional_info = [
        'INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI',
        'BIENES POR CUSTODIO',
        'REGISTRO DE CONSTATACIÓN FÍSICA DE BIENES 2024'
    ]
    start_row_for_info = 2
    for i, info in enumerate(additional_info, start=start_row_for_info):
        ws.append([''] * total_columns)
        info_cell = ws.cell(row=i, column=1)
        info_cell.value = info
        info_cell.font = Font(size=14, bold=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)
    ws.append([''] * total_columns)

    colums_names = ['ID','NombreServicioHTTPS', 'Descripción','IP','Puerto']
    ws.append(colums_names)

    for cell in ws[ws.max_row]:  # Los encabezados están en la fila 3
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    for obj in queryset:
        ws.append([obj.id, obj.NombreServicioHttps,obj.Descripcion, obj.Ip, obj.puerto])
        cell.font = Font(bold=True)
    
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=ws.max_row):  # Empezando desde la fila 4
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    column_widths = [10, 30, 35, 20, 10]  # Ajustar estos valores según tus necesidades
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  # A, B, C, D, E corresponden a 1, 2, 3, 4, 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Apis_Report.xlsx'
    wb.save(response)
    return response




#Formularios ------------------------------------------------------------------------
@login_required
@permission_required('myapp.view_formulario', raise_exception=True)
def formulario_list(request):
    formulario = formulario.objects.all()
    return render(request, 'model_list.html', {'formulario': formulario})

@login_required
@permission_required('myapp.add_formulario', raise_exception=True)
def formulario_create(request):
    if request.method == 'POST':
        form = FormularioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Formulario creado correctamente!')
            return redirect('model_list')
    else:
        form = FormularioForm()
    return render(request, 'formulario/formulario_form.html', {'form': form})

@login_required
@permission_required('myapp.change_formulario', raise_exception=True)
def formulario_update(request, pk):
    formulario = get_object_or_404(Formulario, pk=pk)
    if request.method == 'POST':
        form = FormularioForm(request.POST, instance=formulario)
        if form.is_valid():
            form.save()
            messages.success(request, '!Formulario actualizado correctamente!')
            return redirect('model_list')
    else:
        form = FormularioForm(instance=formulario)
    return render(request, 'formulario/formulario_form.html', {'form': form, 'formulario': formulario})

@login_required
@permission_required('myapp.delete_formulario', raise_exception=True)
def formulario_delete(request, pk):
    formulario = get_object_or_404(Formulario, pk=pk)
    if request.method == 'POST':
        formulario.delete()
        messages.success(request, '¡Formulario eliminado correctamente!')
        return redirect('model_list')
    return render(request, 'formulario/formulario_confirm_delete.html', {'formulario': formulario})


@login_required
@permission_required('myapp.view_formulario', raise_exception=True)
def formulario_detail(request, pk):
    formulario = get_object_or_404(Formulario, pk=pk)
    return render(request, 'formulario/formulario_detail.html', {'formulario': formulario})

def export_to_excel_formulario(request):
    queryset = Formulario.objects.all()

    # Crear un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    # Agregar y formatear el título
    title = 'Reporte de Formulario'
    ws.append([title])
    
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=25, bold=True)  # Negrita y tamaño de fuente
    ws.merge_cells('A1:O1')  # Combinar celdas para el título
    title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto horizontal y verticalmente
    total_columns = 15 
    # Agregar información adicional
    additional_info = [
        'INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI',
        'BIENES POR CUSTODIO',
        'REGISTRO DE CONSTATACIÓN FÍSICA DE BIENES 2024'
    ]
    start_row_for_info = 2
    for i, info in enumerate(additional_info, start=start_row_for_info):
        # Añadir una fila en blanco para que la información adicional se pueda colocar en varias celdas
        ws.append([''] * total_columns)
        info_cell = ws.cell(row=i, column=1)
        info_cell.value = info
        info_cell.font = Font(size=14, bold=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)
    ws.append([''] * total_columns)

       
    # Escribir encabezados de columna
    column_names = ['ID','NombreServidor', 'Ubicacion', 'Marca', 'Modelo', 'NumeroSerie', 'SistemaOperativo', 
                    'PuertoRelevante', 'Entorno','Estado','DireccionIpLocal','DireccionIpPublica','Usuario','Clave','DescripcionProcesos']
    ws.append(column_names)
    
    # Aplicar formato a los encabezados de columna
    for cell in ws[ws.max_row]:  # Los encabezados están en la fila 3
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    # Escribir datos de los objetos en el libro
    for obj in queryset:
        ws.append([obj.id, obj.NombreServidor, obj.Ubicacion, obj.Marca, obj.Modelo, obj.NumeroSerie, obj.SistemaOperativo, 
                   obj.PuertoRelevante, obj.Entorno, obj.Estado, obj.DireccionIpLocal, obj.DireccionIpPublica, obj.Usuario, obj.Clave, obj.DescripcionProcesos])
        cell.font = Font(bold=True)

    # Aplicar formato a las celdas de datos
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=ws.max_row):  # Empezando desde la fila 4
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    column_widths = [10, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 30]  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  

    # Crear una respuesta de HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Formulario_Report.xlsx'
    wb.save(response)
    return response


#Subdominio ----------------------------------------------------------------------------
# Lista de Subdominios
@login_required
@permission_required('myapp.view_subdominios', raise_exception=True)
def subdominios_list(request):
    subdominios = Subdominios.objects.all()
    return render(request, 'model_list.html', {'subdominios': subdominios})

# Crear Subdominio
@login_required
@permission_required('myapp.add_subdominios', raise_exception=True)
def subdominios_create(request):
    if request.method == 'POST':
        form = SubdominioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Subdominio creado correctamente!')
            return redirect('model_list')  # Redirige a la página con la lista de subdominios
    else:
        form = SubdominioForm()
    return render(request, 'subdominios/subdominios_form.html', {'form': form})

# Actualizar Subdominio
@login_required
@permission_required('myapp.change_subdominios', raise_exception=True)
def subdominios_update(request, pk):
    subdominio = get_object_or_404(Subdominios, pk=pk)
    if request.method == 'POST':
        form = SubdominioForm(request.POST, instance=subdominio)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Subdominio actualizado correctamente!')
            return redirect('model_list')  # Redirige a la página con la lista de subdominios
    else:
        form = SubdominioForm(instance=subdominio)
    return render(request, 'subdominios/subdominios_form.html', {'form': form, 'subdominio': subdominio})

# Eliminar Subdominio
@login_required
@permission_required('myapp.delete_subdominios', raise_exception=True)
def subdominios_delete(request, pk):
    subdominio = get_object_or_404(Subdominios, pk=pk)
    if request.method == 'POST':
        subdominio.delete()
        messages.success(request, '¡Subdominio eliminado correctamente!')
        return redirect('model_list')  # Redirige a la página con la lista de subdominios
    return render(request, 'subdominios/subdominios_confirm_delete.html', {'subdominio': subdominio})

# Detalle del Subdominio
@login_required
@permission_required('myapp.view_subdominios', raise_exception=True)
def subdominios_detail(request, pk):
    subdominio = get_object_or_404(Subdominios, pk=pk)
    return render(request, 'subdominios/subdominios_detail.html', {'subdominio': subdominio})

# Exportación a Excel
def export_to_excel_subdominios(request):
    queryset = Subdominios.objects.all()

    # Crear un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    # Agregar y formatear el título
    title = 'Reporte de Subdominios'
    ws.append([title])
    
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=25, bold=True)  # Negrita y tamaño de fuente
    ws.merge_cells('A1:E1')  # Combinar celdas para el título
    title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto horizontal y verticalmente
    total_columns = 5 
    # Agregar información adicional
    additional_info = [
        'INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI',
        'BIENES POR CUSTODIO',
        'REGISTRO DE CONSTATACIÓN FÍSICA DE BIENES 2024'
    ]
    start_row_for_info = 2
    for i, info in enumerate(additional_info, start=start_row_for_info):
        # Añadir una fila en blanco para que la información adicional se pueda colocar en varias celdas
        ws.append([''] * total_columns)
        info_cell = ws.cell(row=i, column=1)
        info_cell.value = info
        info_cell.font = Font(size=14, bold=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_columns)
    ws.append([''] * total_columns)

       
    # Escribir encabezados de columna
    column_names = ['ID', 'Nombre', 'IpPublica', 'IpInterna', 'Host']
    ws.append(column_names)
    
    # Aplicar formato a los encabezados de columna
    for cell in ws[ws.max_row]:  # Los encabezados están en la fila 3
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    # Escribir datos de los objetos en el libro
    for obj in queryset:
        ws.append([obj.Id, obj.Nombre, obj.IpPublica, obj.IpInterna, obj.Host])
        cell.font = Font(bold=True)
        
    # Aplicar formato a las celdas de datos
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=ws.max_row):  # Empezando desde la fila 4
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar texto

    # Ajustar el ancho de las columnas (opcional, según sea necesario)
    column_widths = [10, 30, 20, 20, 30]  # Ajustar estos valores según tus necesidades
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width  # A, B, C, D, E corresponden a 1, 2, 3, 4, 5

    # Crear una respuesta de HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Subdominio_Report.xlsx'
    wb.save(response)
    return response